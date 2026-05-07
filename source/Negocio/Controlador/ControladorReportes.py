from Persistencia.Postgres.ReporteRepositoryImpl import ReporteRepositoryImpl


class ControladorReportes:

    def __init__(self):
        self.repo = ReporteRepositoryImpl()

    def generar_reporte(self, tipo, fecha_inicio=None, fecha_fin=None, tipo_usuario=None):

        # 🔹 normalizar tipo usuario
        if tipo_usuario:
            tipo_usuario = tipo_usuario.upper()
            if tipo_usuario == "ALUMNOS":
                tipo_usuario = "ALUMNO"
            elif tipo_usuario == "VISITANTES":
                tipo_usuario = "VISITANTE"

        if tipo == "Asistencia completa":
            return self.repo.obtener_asistencia_completa(
                fecha_inicio,
                fecha_fin,
                tipo_usuario
            )

        elif tipo == "Usuarios más activos":
            return self.repo.obtener_usuarios_mas_activos(
                fecha_inicio,
                fecha_fin,
                tipo_usuario
            )

        elif tipo == "Horas pico avanzado":
            return self.repo.obtener_horas_pico_avanzado(
                fecha_inicio,
                fecha_fin
            )

        elif tipo == "Tendencia diaria":
            return self.repo.obtener_tendencia_diaria(
                fecha_inicio,
                fecha_fin,
                tipo_usuario
            )

        elif tipo == "Reporte por carrera":
            return self.repo.obtener_reporte_carreras(
                fecha_inicio,
                fecha_fin
            )

        elif tipo == "Reporte por facultad":
            return self.repo.obtener_reporte_facultades(
                fecha_inicio,
                fecha_fin
            )

        elif tipo == "Reporte por semestre":
            return self.repo.obtener_reporte_semestres(
                fecha_inicio,
                fecha_fin
            )

        elif tipo == "Visitantes externos":
            return self.repo.obtener_visitantes_externos(
                fecha_inicio,
                fecha_fin
            )
        
        elif tipo == "Reporte de saturación":
            return self.repo.obtener_reporte_saturacion(
                fecha_inicio,
                fecha_fin
            )

        elif tipo == "Reporte de crecimiento":
            return self.repo.obtener_reporte_crecimiento()

        return []
    
    def exportar_excel(self, ruta, tipo, fecha_inicio=None, fecha_fin=None, tipo_usuario=None):

        import openpyxl
        from openpyxl.utils import get_column_letter

        # obtener datos del reporte
        datos = self.generar_reporte(
            tipo=tipo,
            fecha_inicio=fecha_inicio,
            fecha_fin=fecha_fin,
            tipo_usuario=tipo_usuario
        )

        # validar datos vacíos
        if not datos:
            return False

        # validar formato
        if not isinstance(datos[0], dict):
            raise Exception("El reporte no tiene formato diccionario")

        # crear workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reporte"

        # =========================
        # HEADERS
        # =========================
        headers = list(datos[0].keys())
        ws.append(headers)

        # =========================
        # DATOS
        # =========================
        for d in datos:

            fila = []

            for valor in d.values():

                # convertir fechas automáticamente
                if hasattr(valor, "strftime"):
                    valor = valor.strftime("%Y-%m-%d %H:%M:%S")

                fila.append(valor)

            ws.append(fila)

        # =========================
        # AUTO AJUSTE COLUMNAS
        # =========================
        for col in ws.columns:

            max_length = 0
            col_letter = get_column_letter(col[0].column)

            for cell in col:

                if cell.value is not None:

                    max_length = max(
                        max_length,
                        len(str(cell.value))
                    )

            ws.column_dimensions[col_letter].width = max_length + 2

        # =========================
        # GUARDAR
        # =========================
        wb.save(ruta)

        return True