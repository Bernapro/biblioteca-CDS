from Negocio.Modelo.RepositorioImpl import RepositorioImpl
from Persistencia.CRUD.CRUDimpl import CRUDimp
from Persistencia.Postgres.Pool.DBPool import db
from datetime import datetime, timedelta


class ControladorHistorial:

    def __init__(self):
        self.__repo = RepositorioImpl(CRUDimp())

    def cerrar_registros(self):
        self.__repo.ejecutar_procedimiento("cerrar_registros_pendientes")
    
    def mapear_simple(self, datos):
        """Mapea datos completos a formato simple para tabla"""
        return [
            {
                "identificador": d["identificador"],
                "nombre": d["nombre_completo"],
                "tipo": d["tipo_usuario"].upper(),
                "fecha": d["fecha_entrada"].strftime("%Y-%m-%d"),
                "entrada": d["fecha_entrada"].strftime("%H:%M:%S"),
                "salida": d["fecha_salida"].strftime("%H:%M:%S") if d["fecha_salida"] else "-"
            }
            for d in datos
        ]

    def mapear_completo(self, datos):
        return [
            {
                "id": d["id_registro"],
                "identificador": d["identificador"],
                "nombre_completo": d["nombre_completo"],  # 🔥 clave
                "tipo": d["tipo_usuario"],
                "fecha_entrada": d["fecha_entrada"],
                "fecha_salida": d["fecha_salida"],
                "matricula": d["matricula"],
                "n_plaza": d["n_plaza"],
                "grupo": d["grupo"],
                "carrera": d["nombre_carrera"],
                "facultad": d["nombre_facultad"],
                "semestre": d["semestre"],
                "institucion": d["nombre_institucion"]
            }
            for d in datos
        ]

        # =========================
        # PUBLICOS
        # =========================
    def obtener_historial(
        self,
        texto="",
        fecha_inicio=None,
        fecha_fin=None,
        tipo="Todos",
        estado="Todos",
        limit=10,
        offset=0
    ):

        filtros = {}
        or_filtros = []

        if texto:
            or_filtros = [
                {"nombre_completo__like": texto},
                {"identificador__like": texto}
            ]

        if fecha_inicio:
            if isinstance(fecha_inicio, str):
                fecha_inicio = datetime.strptime(fecha_inicio, "%Y-%m-%d")

            fecha_inicio = fecha_inicio.replace(hour=0, minute=0, second=0)

            filtros["fecha_entrada__gte"] = fecha_inicio

        if fecha_fin:
            if isinstance(fecha_fin, str):
                fecha_fin = datetime.strptime(fecha_fin, "%Y-%m-%d")

            fecha_fin = fecha_fin.replace(hour=23, minute=59, second=59)

            filtros["fecha_entrada__lte"] = fecha_fin

        if tipo != "Todos":
            filtros["tipo_usuario"] = tipo.upper()

        if estado == "Activos":
            filtros["fecha_salida__isnull"] = True
        elif estado == "Finalizados":
            filtros["fecha_salida__notnull"] = True

        datos_raw = self.__repo.obtener_avanzado(
            nombre_tabla="vista_historial",
            filtros=filtros,
            or_filtros=or_filtros,
            order_by=[("fecha_entrada", "DESC")],
            limit=limit,
            offset=offset
        )

        total = self.__repo.contar_avanzado(
            nombre_tabla="vista_historial",
            filtros=filtros,
            or_filtros=or_filtros
        )
        self._datos_completos = self.mapear_completo(datos_raw)
        datos = self.mapear_simple(datos_raw)

        return datos, total
    

    def obtener_historial_completo(
        self,
        texto="",
        fecha_inicio=None,
        fecha_fin=None,
        tipo="Todos",
        estado="Todos"
    ):
        datos, _ = self.obtener_historial(
            texto=texto,
            fecha_inicio=fecha_inicio,
            fecha_fin=fecha_fin,
            tipo=tipo,
            estado=estado,
            limit=None,   
            offset=None
        )

        return self._datos_completos

    def contar_usuarios_hoy(self):
        """Retorna total de usuarios únicos que asistieron hoy
        Procesa datos de vista_historial en memoria (reutilización)
        """
        datos = self.__repo.obtener_avanzado(
            nombre_tabla="vista_historial"
        )
        usuarios_hoy = set()
        
        for d in datos:
            if d["fecha_entrada"].date() == self._get_today():
                usuarios_hoy.add(d["id_usuario"])
        
        return len(usuarios_hoy)

    def _get_today(self):
        """Helper para obtener fecha de hoy"""
        from datetime import date
        return date.today()

    # =========================
    # EXPORTAR EXCEL
    # =========================
    def exportar_excel(self, ruta, texto="", fecha_inicio=None, fecha_fin=None, tipo="Todos", estado="Todos"):
        """Exporta historial filtrado a Excel"""
        if fecha_inicio and isinstance(fecha_inicio, str):
            from datetime import datetime
            fecha_inicio = datetime.strptime(fecha_inicio, "%Y-%m-%d").date()
        if fecha_fin and isinstance(fecha_fin, str):
            from datetime import datetime
            fecha_fin = datetime.strptime(fecha_fin, "%Y-%m-%d").date()

        import openpyxl
        from openpyxl.utils import get_column_letter

        datos = self.obtener_historial_completo(
            texto, fecha_inicio, fecha_fin, tipo, estado
        )

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Historial"

        headers = [
            "ID", "Identificador", "Nombre", "Tipo",
            "Entrada", "Salida",
            "Matrícula", "Plaza",
            "Grupo", "Carrera", "Facultad", "Semestre", "Institución"
        ]

        ws.append(headers)

        for d in datos:
            ws.append([
                d["id"],
                d["identificador"],
                d["nombre_completo"],
                d["tipo"],
                d["fecha_entrada"].strftime("%Y-%m-%d %H:%M:%S"),
                d["fecha_salida"].strftime("%Y-%m-%d %H:%M:%S") if d["fecha_salida"] else "",
                d["matricula"],
                d["n_plaza"],
                d["grupo"],
                d["carrera"],
                d["facultad"],
                d["semestre"],
                d["institucion"],
            ])

        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max_length + 2

        wb.save(ruta)